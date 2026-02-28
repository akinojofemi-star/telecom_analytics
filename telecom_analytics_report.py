# Required packages: pip install pandas matplotlib openpyxl
# How to run: python telecom_analytics_report.py

"""
Telecom Customer Analytics Report Generator

This script analyzes sample telecom customer usage and service data.
It aggregates data, computes KPIs, generates visualizations, and exports
a professional, formatted Excel report to the user's Desktop.
"""

import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as xlImage

def load_sample_data() -> pd.DataFrame:
    """
    Generates realistic sample telecommunications data.
    
    Returns:
        pd.DataFrame: A DataFrame containing 30 rows of customer data.
    """
    data = {
        'Customer_ID': range(1001, 1031),
        'Region': ['Lagos', 'Abuja', 'Port Harcourt', 'Kano', 'Enugu', 'Ibadan', 'Lagos', 'Abuja', 'Lagos', 'Port Harcourt',
                   'Kano', 'Enugu', 'Ibadan', 'Lagos', 'Abuja', 'Lagos', 'Port Harcourt', 'Kano', 'Enugu', 'Ibadan',
                   'Lagos', 'Abuja', 'Lagos', 'Port Harcourt', 'Kano', 'Enugu', 'Ibadan', 'Lagos', 'Abuja', 'Port Harcourt'],
        'Plan_Type': ['Prepaid', 'Postpaid', 'Prepaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Prepaid', 'Postpaid',
                      'Prepaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Prepaid',
                      'Postpaid', 'Prepaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Prepaid', 'Postpaid', 'Prepaid', 'Postpaid', 'Prepaid'],
        'Data_Usage_GB': [15.2, 45.1, 8.5, 12.0, 30.5, 5.2, 55.4, 18.2, 22.1, 40.0,
                          10.5, 9.2, 28.4, 14.5, 38.2, 19.8, 11.2, 35.0, 7.5, 6.8,
                          60.2, 16.5, 25.4, 42.1, 14.0, 8.8, 31.5, 20.1, 48.5, 15.6],
        'Voice_Minutes': [120, 450, 80, 150, 300, 60, 500, 200, 250, 400,
                          100, 90, 280, 140, 380, 210, 110, 350, 75, 65,
                          600, 180, 270, 420, 130, 85, 320, 220, 490, 145],
        'SMS_Count': [50, 200, 20, 40, 150, 15, 250, 80, 100, 180,
                      30, 25, 120, 60, 160, 90, 45, 140, 22, 18,
                      300, 75, 110, 190, 55, 35, 130, 95, 220, 65],
        'Complaint_Category': ['Network', 'None', 'Billing', 'Network', 'Coverage', 'None', 'Device', 'Service', 'None', 'Billing',
                               'Network', 'None', 'Coverage', 'Service', 'None', 'Network', 'None', 'Service', 'Billing', 'None',
                               'Coverage', 'Network', 'None', 'Service', 'Billing', 'None', 'Network', 'None', 'Coverage', 'Billing'],
        'Last_Recharge_Amount_NGN': [2000, 10000, 1000, 1500, 8000, 500, 15000, 2500, 3000, 9000,
                                     1200, 1000, 7500, 2000, 8500, 2500, 1500, 9500, 800, 700,
                                     20000, 2200, 3500, 11000, 1800, 1000, 8000, 2800, 12000, 2000],
        'Report_Month': ['Jan-2026'] * 30
    }
    return pd.DataFrame(data)

def perform_analysis(df: pd.DataFrame) -> dict:
    """
    Analyzes the telecom dataset and computes key business metrics.
    
    Args:
        df: The telecom customer dataset.
        
    Returns:
        A dictionary containing computed aggregations and KPIs.
    """
    results = {}
    
    # 1. Overall KPIs
    results['total_customers'] = len(df)
    results['prepaid_pct'] = (df['Plan_Type'] == 'Prepaid').mean() * 100
    results['postpaid_pct'] = (df['Plan_Type'] == 'Postpaid').mean() * 100
    
    results['avg_data'] = df['Data_Usage_GB'].mean()
    results['avg_voice'] = df['Voice_Minutes'].mean()
    results['avg_sms'] = df['SMS_Count'].mean()
    
    # 2. Aggregations by Region and Plan_Type
    results['region_agg'] = df.groupby('Region')[['Data_Usage_GB', 'Voice_Minutes', 'SMS_Count']].mean().reset_index()
    results['plan_agg'] = df.groupby('Plan_Type')[['Data_Usage_GB', 'Voice_Minutes', 'SMS_Count']].mean().reset_index()
    
    # 3. Complaints Analysis
    complaints = df[df['Complaint_Category'] != 'None']
    
    complaint_counts = complaints['Complaint_Category'].value_counts().reset_index()
    complaint_counts.columns = ['Complaint_Category', 'Count']
    complaint_counts['Percentage'] = (complaint_counts['Count'] / len(complaints)) * 100
    results['complaint_category_dist'] = complaint_counts
    
    if not complaint_counts.empty:
        results['top_complaint_cat'] = complaint_counts.iloc[0]['Complaint_Category']
    else:
        results['top_complaint_cat'] = 'None'
    
    # Complaint count by region and category
    results['complaint_region_cat'] = pd.crosstab(df['Region'], df['Complaint_Category']).reset_index()
    
    # Region with highest complaint volume relative to customer count
    region_complaints = complaints.groupby('Region').size().reset_index(name='Complaint_Count')
    region_customers = df.groupby('Region').size().reset_index(name='Customer_Count')
    ratio_df = pd.merge(region_customers, region_complaints, on='Region', how='left').fillna(0)
    ratio_df['Complaint_Rate'] = ratio_df['Complaint_Count'] / ratio_df['Customer_Count']
    results['complaint_rates'] = ratio_df.sort_values('Complaint_Rate', ascending=False)
    
    # 4. Top 5 Regions by Avg Data Usage
    results['top_regions_data'] = results['region_agg'].sort_values('Data_Usage_GB', ascending=False).head(5)
    
    return results

def generate_visualizations(df: pd.DataFrame, analysis: dict):
    """
    Creates and saves analytical visualizations as high-quality PNG files temporarily.
    """
    colors = ['#4C72B0', '#DD8452', '#55A868', '#C44E52', '#8172B3', '#937860']
    
    # 1. Horizontal bar chart: Average Data Usage by Region
    plt.figure(figsize=(8, 5))
    top_data = analysis['top_regions_data'].sort_values('Data_Usage_GB', ascending=True)
    plt.barh(top_data['Region'], top_data['Data_Usage_GB'], color='#4C72B0')
    plt.title('Top 5 Regions by Average Data Usage (GB)', fontsize=14, pad=15)
    plt.xlabel('Average Data Usage (GB)')
    plt.ylabel('Region')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)
    plt.tight_layout()
    plt.savefig('data_usage_bar.png', dpi=300)
    plt.close()
    
    # 2. Pie chart: Complaint Category Distribution
    complaints = df[df['Complaint_Category'] != 'None']
    cat_counts = complaints['Complaint_Category'].value_counts()
    
    plt.figure(figsize=(7, 6))
    plt.pie(cat_counts, labels=cat_counts.index, autopct='%1.1f%%', startangle=140, 
            colors=colors[:len(cat_counts)], wedgeprops={'edgecolor': 'white'})
    plt.title('Complaint Category Distribution (%)', fontsize=14, pad=15)
    plt.tight_layout()
    plt.savefig('complaint_pie.png', dpi=300)
    plt.close()
    
    # 3. Stacked bar chart: Complaint counts by Region (Top 3 Categories)
    top_3_cats = cat_counts.nlargest(3).index.tolist()
    stacked_data = df[df['Complaint_Category'].isin(top_3_cats)]
    cross_tab = pd.crosstab(stacked_data['Region'], stacked_data['Complaint_Category'])
    
    fig, ax = plt.subplots(figsize=(10, 6))
    # Fill colors to make it professional
    cross_tab.plot(kind='bar', stacked=True, ax=ax, color=colors[:len(top_3_cats)])
    plt.title('Complaint Counts by Region (Top 3 Categories)', fontsize=14, pad=15)
    plt.xlabel('Region')
    plt.ylabel('Number of Complaints')
    plt.legend(title='Complaint Category', frameon=False, loc='upper right')
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.gca().spines['top'].set_visible(False)
    plt.gca().spines['right'].set_visible(False)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('complaint_stacked.png', dpi=300)
    plt.close()

def apply_header_style(ws, freeze_panes=True):
    """Applies standardized styling to the header row of a worksheet."""
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    if freeze_panes:
        ws.freeze_panes = 'A2'
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def create_excel_report(df: pd.DataFrame, analysis: dict, filepath: str):
    """
    Builds the multi-sheet Excel report, formatting and saving it to the specified filepath.
    """
    wb = Workbook()
    wb.remove(wb.active) # Remove default 'Sheet'
    
    # Sheet 1: Executive Summary
    ws_exec = wb.create_sheet('Executive Summary')
    ws_exec.column_dimensions['A'].width = 30
    ws_exec.column_dimensions['B'].width = 20
    ws_exec.column_dimensions['D'].width = 65
    
    ws_exec['A1'] = 'Telecom Customer Analytics - Executive Summary'
    ws_exec['A1'].font = Font(size=16, bold=True, color='1F4E78')
    ws_exec.merge_cells('A1:B1')
    
    # KPIs definitions
    kpis = [
        ('Total Customers Analyzed', analysis['total_customers']),
        ('Prepaid Customers (%)', f"{analysis['prepaid_pct']:.1f}%"),
        ('Postpaid Customers (%)', f"{analysis['postpaid_pct']:.1f}%"),
        ('Average Data Usage (GB)', f"{analysis['avg_data']:.2f}"),
        ('Average Voice Minutes', f"{int(analysis['avg_voice'])}"),
        ('Average SMS Count', f"{int(analysis['avg_sms'])}"),
        ('Top Complaint Category', analysis['top_complaint_cat'])
    ]
    
    ws_exec['A3'] = 'Key Performance Indicators'
    ws_exec['A3'].font = Font(bold=True)
    ws_exec['A3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for i, (metric, value) in enumerate(kpis, start=4):
        ws_exec.cell(row=i, column=1, value=metric)
        cell_val = ws_exec.cell(row=i, column=2, value=value)
        cell_val.alignment = Alignment(horizontal='right')
    
    # Business recommendations text
    ws_exec['D3'] = 'Business Recommendations'
    ws_exec['D3'].font = Font(bold=True)
    ws_exec['D3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    recommendations = [
        "1. Network infrastructure in high-usage regions like Lagos and Abuja requires strategic capacity upgrades.",
        "2. Prepaid customers represent the vast majority; marketing should focus on flexible data bundles.",
        "3. The high proportion of Network-related complaints highlights an immediate need for localized cell tower maintenance.",
        "4. Introduce loyalty rewards for Postpaid users to encourage plan upgrades and increase overall average revenue.",
        "5. Regions displaying high complaints relative to their customer bases should undergo rigorous service audits.",
        "6. Proactively target high-data consumers with tailored service tiers to better monetize their data usage trends.",
        "7. Reduce recurring Billing-related complaints by significantly improving self-service application transparency."
    ]
    
    for i, rec in enumerate(recommendations, start=4):
        ws_exec.cell(row=i, column=4, value=rec)
    
    # Sheet 2: Sample Data
    ws_data = wb.create_sheet('Sample Data')
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    apply_header_style(ws_data)
    
    for row in ws_data.iter_rows(min_row=2, max_col=ws_data.max_column):
        row[3].number_format = '0.00' # Data_Usage_GB
        row[7].number_format = '#,##0.00' # Last_Recharge_Amount_NGN
    
    # Sheet 3: Regional Summary
    ws_region = wb.create_sheet('Regional Summary')
    for r in dataframe_to_rows(analysis['region_agg'], index=False, header=True):
        ws_region.append(r)
    apply_header_style(ws_region)
    for row in ws_region.iter_rows(min_row=2, max_col=ws_region.max_column):
        row[1].number_format = '0.00' 
        row[2].number_format = '#,##0'
        row[3].number_format = '#,##0'
    
    # Sheet 4: Plan Type Summary
    ws_plan = wb.create_sheet('Plan Type Summary')
    for r in dataframe_to_rows(analysis['plan_agg'], index=False, header=True):
        ws_plan.append(r)
    apply_header_style(ws_plan)
    for row in ws_plan.iter_rows(min_row=2, max_col=ws_plan.max_column):
        row[1].number_format = '0.00'
        row[2].number_format = '#,##0'
        row[3].number_format = '#,##0'

    # Sheet 5: Complaints Analysis
    ws_comp = wb.create_sheet('Complaints Analysis')
    
    ws_comp['A1'] = 'Complaint Distribution by Category'
    ws_comp['A1'].font = Font(bold=True)
    ws_comp.append([]) # space to avoid overwrite
    for r in dataframe_to_rows(analysis['complaint_category_dist'], index=False, header=True):
        ws_comp.append(r)
        
    start_row = ws_comp.max_row + 3
    ws_comp.cell(row=start_row, column=1, value='Complaint Frequency Relative to Customer Count')
    ws_comp.cell(row=start_row, column=1).font = Font(bold=True)
    
    # Add table headers manually for spacing
    ratio_df = analysis['complaint_rates']
    for r in dataframe_to_rows(ratio_df, index=False, header=True):
        ws_comp.append(r)
        
    # Styling inner headers explicitly
    for cell in ws_comp[2]:
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        
    for cell in ws_comp[start_row + 1]:
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Format specific rows for %
    for row in ws_comp.iter_rows(min_row=3, max_row=start_row-3, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '0.00'
            
    for row in ws_comp.iter_rows(min_row=start_row + 2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '0.00%'
            
    ws_comp.column_dimensions['A'].width = 20
    ws_comp.column_dimensions['B'].width = 15
    ws_comp.column_dimensions['C'].width = 20
    ws_comp.column_dimensions['D'].width = 20
    ws_comp.column_dimensions['E'].width = 20

    # Sheet 6: Visualizations
    ws_viz = wb.create_sheet('Visualizations')
    ws_viz.column_dimensions['A'].width = 2
    
    # Embed the saved PNG files
    if os.path.exists('data_usage_bar.png'):
        img1 = xlImage('data_usage_bar.png')
        ws_viz.add_image(img1, 'B2')
        
    if os.path.exists('complaint_pie.png'):
        img2 = xlImage('complaint_pie.png')
        ws_viz.add_image(img2, 'B30')
        
    if os.path.exists('complaint_stacked.png'):
        img3 = xlImage('complaint_stacked.png')
        ws_viz.add_image(img3, 'N2')

    # Save logic with structured error handling
    try:
        wb.save(filepath)
        print(f"Analysis complete. Professional report saved to Desktop as {os.path.basename(filepath)}")
    except PermissionError:
        print(f"Error: Permission denied. Please close the file {os.path.basename(filepath)} if it is currently open in another program and try again.")
    except Exception as e:
        print(f"Error saving report: {str(e)}")

def main():
    """Main execution function defining the script workflow."""
    # Define Desktop output path (cross-platform compatible)
    desktop_path = os.path.expanduser("~/Desktop")
    report_filename = "Telecom_Customer_Analytics_Report_2026.xlsx"
    filepath = os.path.join(desktop_path, report_filename)
    
    # Execute analytical pipeline
    df = load_sample_data()
    analysis = perform_analysis(df)
    generate_visualizations(df, analysis)
    create_excel_report(df, analysis, filepath)
    
    # Clean up temporary chart images
    for img in ['data_usage_bar.png', 'complaint_pie.png', 'complaint_stacked.png']:
        if os.path.exists(img):
            try:
                os.remove(img)
            except OSError:
                pass

if __name__ == "__main__":
    main()
